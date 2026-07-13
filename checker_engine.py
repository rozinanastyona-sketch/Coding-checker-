from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
import json
import re
import shutil

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.comments import Comment as CellComment
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


@dataclass
class EventRow:
    original_index: int
    data: Dict[str, Any]
    behavior: str
    modifiers: List[str]
    comment: str


@dataclass
class Utterance:
    """
    A single coded utterance: one "Communicative Intent Present/Absent" event
    plus all the coding rows and start/stop markers grouped around it (see
    segment_utterances()).

    Transcription conventions inside the Comment cell on the CI row
    (per the AAC Lab Operational Definitions doc, "Utterance Boundaries &
    Transcription Conventions" / "Number of Relevant Symbols" sections):
      - Square brackets `[...]` mark symbols the coder judged IRRELEVANT
        (e.g. "DOG [IN ON] UNDER CAR" -> DOG, UNDER, CAR are relevant; IN, ON
        are not). This is a human judgment call (mazes, non-adjacent repeats,
        context-dependent relevance - see "Number of Relevant Symbols" in the
        Operational Definitions) and is NOT mechanically re-derivable from
        the bracketed text alone, so this engine does not try to compute
        Number of Symbols / Number of Relevant Symbols from the transcript.
        Those are taken directly from the "Number of Symbols" / "Number of
        Relevant Symbols" behavior rows the coder entered, same as every
        other coded value, and compared against the key the same way.
      - A leading dash (e.g. "-S", "-ING", "-'S") marks a bound morpheme
        attached to the previous symbol, not a separate word.
      - Parentheses are sometimes used for a coder's own aside/clarification
        (e.g. "HIPPO -S (BARN)").
    None of the above characters are stripped or specially parsed anywhere in
    this file - comment text is always treated as opaque display text
    (see clean_text/utterance_text below) specifically so that bracket/dash
    notation always survives unchanged into Review_Notes and the "Utterance
    #" transcript shown to the reviewer.
    """
    uid: int
    rows: List[EventRow] = field(default_factory=list)
    boundary_unclear: bool = False

    @property
    def first_row_index(self) -> int:
        return self.rows[0].original_index if self.rows else -1

    @property
    def anchor_row_index(self) -> int:
        """Row to attach an utterance-level flag to.

        first_row_index is often a stray "utterance start" marker (Observer
        duplicates them, and coders sometimes retype the transcript on one),
        which makes the highlight land on a meaningless row. Prefer the
        Communicative Intent row, then the first coded row.
        """
        for row in self.rows:
            if behavior_in(row.behavior, ("Communicative Intent Present", "Communicative Intent Absent")):
                return row.original_index
        for row in self.rows:
            b = behavior_key(row.behavior)
            if not b.startswith("utterance start") and not b.startswith("utterance stop"):
                return row.original_index
        return self.first_row_index

    @property
    def last_row_index(self) -> int:
        return self.rows[-1].original_index if self.rows else -1

    @property
    def utterance_text(self) -> str:
        return self._text_info()[0]

    @property
    def text_is_partial(self) -> bool:
        """True when the transcript was recovered from a 'USV: ...' note.

        A USV note records only the subject-verb combination (e.g.
        "USV: I PUT"), not the full utterance, so such text is a prefix /
        subset of the real transcript and must be matched by containment,
        not full-string similarity, during alignment.
        """
        return self._text_info()[1]

    def _text_info(self) -> Tuple[str, bool]:
        usv_prefix = "USV:"

        def extract(c: str) -> Optional[Tuple[str, bool]]:
            if not c:
                return None
            if c.upper().startswith(usv_prefix):
                # Some coders combine the USV note and the transcript in one
                # cell (e.g. "USV: I PUT") instead of putting the note on the
                # SV row and the transcript on the CI row. Strip the prefix
                # and keep whatever transcript remains.
                remainder = c[len(usv_prefix):].strip()
                return (remainder, True) if remainder else None
            return (c, False)

        # Preferred source: the Communicative Intent row. Absent counts too -
        # a coder can judge that no communicative intent was present and still
        # transcribe what was produced.
        for row in self.rows:
            if behavior_in(row.behavior, ("Communicative Intent Present", "Communicative Intent Absent")):
                got = extract(clean_text(row.comment))
                if got:
                    return got

        # Fallback: the CI row is missing entirely (real key files contain
        # utterances coded without one). The transcript then sits in the
        # Comment of the next coding row. Start/stop markers are skipped
        # explicitly: coders sometimes retype the previous transcript on a
        # stray marker row, which would otherwise be picked up as this
        # utterance's text.
        for row in self.rows:
            b = behavior_key(row.behavior)
            if b.startswith("utterance start") or b.startswith("utterance stop"):
                continue
            got = extract(clean_text(row.comment))
            if got:
                return got
        return "", False


@dataclass
class Issue:
    kind: str
    color: str
    message: str
    row_index: Optional[int] = None
    insert_after_row_index: Optional[int] = None
    missing_behavior: Optional[str] = None
    expected: Optional[str] = None
    utterance_id: Optional[int] = None


def load_grammar(path: str | Path) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        grammar = yaml.safe_load(f)
    # Student-facing wording lives in feedback.yaml next to grammar.yaml, so the
    # lab can reword feedback without touching either the code or the coding scheme.
    fb_path = Path(path).with_name("feedback.yaml")
    if fb_path.exists():
        with open(fb_path, "r", encoding="utf-8") as f:
            grammar["feedback"] = yaml.safe_load(f) or {}
    else:
        grammar["feedback"] = {}
    return grammar


def _fmt(value: Any, decimals: bool = False) -> str:
    """Render a modifier value for a student.

    Counts read better as whole numbers (4.0 -> 4). Word Order scores use
    halves (0.0 / 0.5 / 1.0) and are passed decimals=True, so "you coded: 0.0"
    doesn't read like a truncated sentence.
    """
    s = clean_text(value)
    if not s:
        return "-"
    try:
        f = float(s)
    except ValueError:
        return s
    if f != int(f):
        return str(f)
    return f"{f:.1f}" if decimals else str(int(f))


# Categories whose values are scores rather than counts, so 0 -> "0.0".
DECIMAL_CATEGORIES = {"Word Order"}


def _join(values: Iterable[Any], decimals: bool = False) -> str:
    vals = [_fmt(v, decimals) for v in values if clean_text(v)]
    return ", ".join(vals) if vals else "-"


def short_label(grammar: Dict[str, Any], value: str) -> str:
    """Shorten a long official behavior name for use inside a comment."""
    labels = grammar.get("feedback", {}).get("value_labels", {}) or {}
    for full, short in labels.items():
        if behavior_equals(full, value):
            return str(short)
    return value


def render_feedback(
    grammar: Dict[str, Any],
    category: Optional[str],
    situation: str,
    utterance: str = "",
    key: str = "",
    found: str = "",
    missing: str = "",
    extra: str = "",
) -> str:
    """Build the student-facing comment for one issue from feedback.yaml."""
    fb = grammar.get("feedback", {})
    template = None
    if category:
        template = (fb.get("categories", {}).get(category, {}) or {}).get(situation)
    if template is None:
        template = (fb.get("other", {}) or {}).get(situation)
    if template is None:
        template = (fb.get("fallback", {}) or {}).get(situation)
    if template is None:
        template = "{category} - key: {key}, you coded: {found}."

    text = str(template).format(
        category=category or "",
        utterance=utterance,
        key=key or "-",
        found=found or "-",
        missing=missing or "-",
        extra=extra or "-",
    )
    return " ".join(text.split())


def conditional_modifier_note(grammar: Dict[str, Any], modifier: str, utterance: str) -> Optional[str]:
    """Special explanation for modifiers that only apply under a condition."""
    table = grammar.get("feedback", {}).get("conditional_modifiers", {}) or {}
    for name, cfg in table.items():
        if behavior_equals(name, modifier):
            text = str((cfg or {}).get("explanation", "")).format(utterance=utterance)
            return " ".join(text.split()) or None
    return None


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    return str(value).strip()


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def build_alias_map(grammar: Dict[str, Any]) -> Dict[str, str]:
    alias_to_canonical = {}
    for canonical, aliases in grammar.get("aliases", {}).items():
        alias_to_canonical[norm(canonical).lower()] = canonical
        for alias in aliases or []:
            alias_to_canonical[norm(alias).lower()] = canonical
    return alias_to_canonical


def canonicalize(value: Any, alias_map: Dict[str, str]) -> str:
    v = norm(value)
    return alias_map.get(v.lower(), v)


def behavior_key(value: Any) -> str:
    """Case/whitespace-insensitive key for comparing Behavior values.

    Observer XT exports are not consistent about casing across an ethogram
    (e.g. real exports contain "utterance start" but also "Word Order Score"
    and "Parts of Speech Present" with different capitalization than what
    was originally typed into grammar.yaml). Every place that used to compare
    behaviors with `==` or `in` against grammar-defined names now goes
    through behavior_equals()/behavior_in() instead, so a future casing
    mismatch degrades gracefully instead of silently breaking matching.
    """
    return norm(value).lower()


def behavior_equals(a: Any, b: Any) -> bool:
    return behavior_key(a) == behavior_key(b)


def behavior_in(value: Any, targets: Iterable[Any]) -> bool:
    v = behavior_key(value)
    return any(v == behavior_key(t) for t in targets)


def read_excel_first_sheet_or_named(path_or_file: Any, sheet_name: Optional[str] = None) -> pd.DataFrame:
    if sheet_name is None:
        return pd.read_excel(path_or_file, sheet_name=0)
    return pd.read_excel(path_or_file, sheet_name=sheet_name)


def prepare_dataframe(df: pd.DataFrame, grammar: Dict[str, Any]) -> pd.DataFrame:
    """Keep required columns + all Modifier* columns. Do not assume how many modifiers exist."""
    df = df.copy()
    required = grammar["columns"]["required"]
    modifier_prefix = grammar["columns"]["modifier_columns"]["detect_by_prefix"]

    # Observer files may contain whitespace in column names.
    df.columns = [str(c).strip() for c in df.columns]

    missing_required = [c for c in required if c not in df.columns]
    if missing_required:
        raise ValueError(f"Missing required columns: {missing_required}")

    modifier_cols = [c for c in df.columns if str(c).strip().startswith(modifier_prefix)]
    keep_cols = required + modifier_cols
    return df[keep_cols].copy()


def row_to_event(row: pd.Series, idx: int, grammar: Dict[str, Any], alias_map: Dict[str, str]) -> EventRow:
    modifier_prefix = grammar["columns"]["modifier_columns"]["detect_by_prefix"]
    modifier_cols = [c for c in row.index if str(c).startswith(modifier_prefix)]
    ignored = {norm(m).lower() for m in grammar["columns"]["modifier_columns"].get("ignored_modifiers", []) or []}
    modifiers = []
    for col in modifier_cols:
        val = clean_text(row.get(col))
        if not val:
            continue
        canon = canonicalize(val, alias_map)
        if norm(canon).lower() in ignored:
            # Not part of the graded coding scheme (see grammar.yaml
            # ignored_modifiers). Dropped here so it can never produce a
            # "Missing/Extra modifiers" issue or affect the Scores sheet,
            # regardless of whether a given key file still contains it.
            continue
        modifiers.append(canon)

    behavior = canonicalize(row.get("Behavior"), alias_map)
    comment = clean_text(row.get("Comment"))
    data = {str(k): row.get(k) for k in row.index}
    data["Behavior"] = behavior
    data["Comment"] = comment
    return EventRow(original_index=idx, data=data, behavior=behavior, modifiers=modifiers, comment=comment)


def dataframe_to_events(df: pd.DataFrame, grammar: Dict[str, Any]) -> List[EventRow]:
    alias_map = build_alias_map(grammar)
    return [row_to_event(row, idx, grammar, alias_map) for idx, row in df.iterrows()]


def is_ignored_comment(comment: str, grammar: Dict[str, Any]) -> bool:
    prefixes = grammar["utterance_detection"].get("ignore_comment_prefixes", [])
    c = clean_text(comment)
    return any(c.upper().startswith(p.upper()) for p in prefixes)


def segment_utterances(events: List[EventRow], grammar: Dict[str, Any]) -> List[Utterance]:
    """
    Segment the event stream into utterances, anchored on the Listing row.

    No boundary marker in an Observer XT export is reliable on its own; every
    candidate anchor is broken in at least one real file:

      file            utterances   CI rows   Listing   Utterance start
      KEY_Video_1         47          45        47          47
      KEY_Video_2         34          34        34          34
      KEY_Video_3         32          32        32          32
      KEY_Video_4         34          34        34          30
      Carley student      34          34        34          70   (!)

    - "Utterance start" is duplicated or dropped (70 markers for 34
      utterances in one student file; 30 for 34 in KEY_Video_4).
      grammar.yaml's `duplicate_start_stop` note already calls the duplicates
      "Observer app bug, not student error".
    - The Communicative Intent row is sometimes omitted entirely
      (KEY_Video_1: two utterances have no CI row, the transcript sits in the
      Comment of the following coding row instead).
    - The Listing row, however, appears exactly once per utterance in every
      file above, so it is used as the anchor.

    Each utterance therefore runs from its Listing row back through the
    header rows that belong to it (its Communicative Intent row plus any
    start/stop markers) and forward through its remaining coding rows.
    Communicative Intent is kept as a backup anchor in case a coder omits a
    Listing row, and an utterance with no CI row is flagged boundary_unclear
    so the reviewer knows the segmentation there is inferred.
    """
    intents = grammar["utterance_detection"]["intent_behaviors"]
    listing_entry = get_sequence_entry(grammar, "Listing is Present")
    listing_behaviors = listing_entry["behaviors"] if listing_entry else []

    def is_listing(ev: EventRow) -> bool:
        return behavior_in(ev.behavior, listing_behaviors)

    def is_ci(ev: EventRow) -> bool:
        return behavior_in(ev.behavior, intents)

    def is_header(ev: EventRow) -> bool:
        """Rows that belong to the NEXT utterance if they trail the current one."""
        return is_ci(ev) or get_sequence_entry(grammar, ev.behavior) is None

    utterances: List[Utterance] = []
    current: List[EventRow] = []
    uid = 1

    def flush(rows: List[EventRow]) -> None:
        nonlocal uid
        if not rows:
            return
        if not any(get_sequence_entry(grammar, ev.behavior) for ev in rows):
            # Stray markers only: never their own utterance.
            if utterances:
                utterances[-1].rows.extend(rows)
                return
        utt = Utterance(uid=uid, rows=rows)
        uid += 1
        if not any(is_ci(ev) for ev in utt.rows):
            utt.boundary_unclear = True
        utterances.append(utt)

    def split_off_header() -> List[EventRow]:
        """Peel the trailing header rows (start/stop markers + CI) off `current`."""
        nonlocal current
        i = len(current)
        while i > 0 and is_header(current[i - 1]):
            i -= 1
        header = current[i:]
        current = current[:i]
        return header

    for ev in events:
        new_utterance = (
            (is_listing(ev) and any(is_listing(e) for e in current))
            or (is_ci(ev) and any(is_ci(e) for e in current))
        )
        if new_utterance:
            header = split_off_header()
            flush(current)
            current = header
        current.append(ev)

    flush(current)
    return utterances


def get_sequence_entry(grammar: Dict[str, Any], behavior: str) -> Optional[Dict[str, Any]]:
    for entry in grammar["sequence"]:
        if behavior_in(behavior, entry.get("behaviors", [])):
            return entry
    return None


def comparable_rows(utt: Utterance, grammar: Dict[str, Any]) -> List[EventRow]:
    start_b = grammar["utterance_detection"]["start_behavior"]
    stop_b = grammar["utterance_detection"]["stop_behavior"]
    return [r for r in utt.rows if not behavior_in(r.behavior, (start_b, stop_b))]


def find_row_by_group(utt: Utterance, entry: Dict[str, Any], grammar: Dict[str, Any]) -> Optional[EventRow]:
    behaviors = entry.get("behaviors", [])
    for row in comparable_rows(utt, grammar):
        if behavior_in(row.behavior, behaviors):
            return row
    return None


def modifier_set(row: Optional[EventRow]) -> set[str]:
    if row is None:
        return set()
    return {norm(m) for m in row.modifiers if norm(m)}


def validate_row_against_key(
    student_row: EventRow,
    key_row: EventRow,
    entry: Dict[str, Any],
    grammar: Dict[str, Any],
    utterance_text: str = "",
    key_text: str = "",
) -> Optional[str]:
    """Return the student-facing comment for this row, or None if it matches the key.

    When the student's transcript itself differs from the key's, a category may
    supply a "<situation>_transcript_differs" template: e.g. a wrong symbol count
    usually needs no explanation, but if the utterance was transcribed wrongly in
    the first place, the count follows from that and the student should be told so.
    """
    category = entry.get("group")
    transcript_differs = bool(
        utterance_text and key_text and _match_text(utterance_text) != _match_text(key_text)
    )

    def feedback(situation: str, **kw) -> str:
        if transcript_differs:
            variant = render_feedback(grammar, category, situation + "_transcript_differs", **kw)
            # render_feedback falls back to the generic template when the variant is
            # absent; detect that by checking the category actually defines it.
            defined = (grammar.get("feedback", {}).get("categories", {})
                       .get(category, {}) or {}).get(situation + "_transcript_differs")
            if defined:
                return variant
        return render_feedback(grammar, category, situation, **kw)

    if not behavior_equals(student_row.behavior, key_row.behavior):
        return feedback(
            "wrong_value",
            utterance=utterance_text,
            key=short_label(grammar, key_row.behavior),
            found=short_label(grammar, student_row.behavior),
        )

    expected_mods = modifier_set(key_row)
    student_mods = modifier_set(student_row)
    if expected_mods == student_mods:
        return None

    dec = category in DECIMAL_CATEGORIES

    missing = sorted(expected_mods - student_mods)
    extra = sorted(student_mods - expected_mods)

    # A modifier that only applies under a condition (e.g. Subject-Verb Agreement
    # needs a 'to be' verb) gets its own explanation rather than a bare "extra".
    notes = []
    unexplained_extra = []
    for m in extra:
        note = conditional_modifier_note(grammar, m, utterance_text)
        if note:
            notes.append(note)
        else:
            unexplained_extra.append(m)

    if missing:
        notes.insert(0, feedback(
            "missing_mods",
            utterance=utterance_text,
            key=_join(sorted(expected_mods), dec),
            found=_join(sorted(student_mods), dec),
            missing=_join(missing, dec),
            extra=_join(unexplained_extra, dec),
        ))
    elif unexplained_extra:
        notes.insert(0, feedback(
            "extra_mods",
            utterance=utterance_text,
            key=_join(sorted(expected_mods), dec),
            found=_join(sorted(student_mods), dec),
            missing=_join(missing, dec),
            extra=_join(unexplained_extra, dec),
        ))

    return " ".join(notes) if notes else None


def check_usv_rule(student_utt: Utterance, grammar: Dict[str, Any]) -> Optional[Issue]:
    for rule in grammar.get("special_rules", []):
        if rule.get("name") != "USV comment required":
            continue
        for row in comparable_rows(student_utt, grammar):
            if not behavior_equals(row.behavior, rule.get("if_behavior")):
                continue
            mods = modifier_set(row)
            required = set(rule.get("required_modifiers", []))
            if required.issubset(mods):
                prefix = rule.get("then_comment_must_start_with", "USV:")
                # USV note may be in the same row or in any row inside the same utterance.
                has_usv = any(clean_text(r.comment).upper().startswith(prefix.upper()) for r in student_utt.rows)
                if not has_usv:
                    return Issue(
                        kind="usv_missing_note",
                        color=grammar["colors"]["usv_missing_note"],
                        message=render_feedback(grammar, None, "usv_missing_note",
                                                utterance=student_utt.utterance_text),
                        row_index=row.original_index,
                        utterance_id=student_utt.uid,
                    )
    return None


def compare_utterances(student_utt: Utterance, key_utt: Utterance, grammar: Dict[str, Any]) -> List[Issue]:
    issues: List[Issue] = []
    red = grammar["colors"]["error"]
    purple = grammar["colors"]["boundary_unclear"]

    if student_utt.boundary_unclear:
        issues.append(Issue(
            kind="boundary_unclear",
            color=purple,
            message=render_feedback(grammar, None, "boundary_unclear",
                                    utterance=student_utt.utterance_text),
            row_index=student_utt.anchor_row_index,
            utterance_id=student_utt.uid,
        ))

    previous_student_row: Optional[EventRow] = None
    utterance_text = student_utt.utterance_text or key_utt.utterance_text
    for entry in grammar["sequence"]:
        key_row = find_row_by_group(key_utt, entry, grammar)
        student_row = find_row_by_group(student_utt, entry, grammar)

        if key_row is None:
            # Reference does not require this group for this utterance.
            continue

        if student_row is None:
            insert_after = previous_student_row.original_index if previous_student_row else student_utt.first_row_index
            expected = key_row.behavior
            if key_row.modifiers:
                expected += " | " + " | ".join(key_row.modifiers)
            key_desc = short_label(grammar, key_row.behavior)
            if key_row.modifiers:
                key_desc += " (" + _join(key_row.modifiers) + ")"
            issues.append(Issue(
                kind="missing",
                color=red,
                message=render_feedback(
                    grammar, entry.get("group"), "missing_row",
                    utterance=utterance_text, key=key_desc,
                ),
                insert_after_row_index=insert_after,
                missing_behavior=key_row.behavior,
                expected=expected,
                utterance_id=student_utt.uid,
            ))
            continue

        msg = validate_row_against_key(
            student_row, key_row, entry, grammar,
            utterance_text=student_utt.utterance_text or utterance_text,
            key_text=key_utt.utterance_text,
        )
        if msg:
            issues.append(Issue(
                kind="wrong",
                color=red,
                message=msg,
                row_index=student_row.original_index,
                utterance_id=student_utt.uid,
            ))

        previous_student_row = student_row

    # Extra rows: student has known code groups that key does not have in that utterance.
    for srow in comparable_rows(student_utt, grammar):
        entry = get_sequence_entry(grammar, srow.behavior)
        if entry is None:
            continue
        krow = find_row_by_group(key_utt, entry, grammar)
        if krow is None:
            issues.append(Issue(
                kind="extra",
                color=red,
                message=render_feedback(grammar, None, "extra_code",
                                        utterance=student_utt.utterance_text,
                                        found=srow.behavior),
                row_index=srow.original_index,
                utterance_id=student_utt.uid,
            ))

    usv_issue = check_usv_rule(student_utt, grammar)
    if usv_issue:
        issues.append(usv_issue)

    return issues


def _match_text(s: str) -> str:
    """Normalize an utterance transcript for alignment matching only.

    Brackets/parentheses are stripped here because the coder's
    relevance-notation may legitimately differ between student and key
    ("I BUY [SHAKE]" vs "I BUY SHAKE" is the same utterance); this
    normalization is used ONLY to decide which utterances correspond, never
    for scoring or display.
    """
    t = re.sub(r"[\[\]\(\)]", " ", s or "")
    t = re.sub(r"\s+", " ", t).strip().upper()
    return t


def _text_similarity(a: str, b: str, a_partial: bool = False, b_partial: bool = False) -> float:
    from difflib import SequenceMatcher
    na, nb = _match_text(a), _match_text(b)
    if not na and not nb:
        return 0.5  # both untranscribed: weak evidence, allow positional matching
    if not na or not nb:
        return 0.0
    ratio = SequenceMatcher(None, na, nb).ratio()
    # A 'USV: ...' note records only the subject-verb combination, not the
    # full transcript. If the partial side's words all appear (in order) in
    # the other side's transcript, treat that as strong evidence of a match.
    if a_partial or b_partial:
        short, long_ = (na, nb) if a_partial else (nb, na)
        short_words = short.split()
        long_words = long_.split()
        it = iter(long_words)
        contained = all(w in it for w in short_words)
        if contained:
            ratio = max(ratio, 0.7)
    return ratio


def align_utterances(
    student_utts: List[Utterance],
    key_utts: List[Utterance],
    min_similarity: float = 0.6,
) -> List[Tuple[Optional[int], Optional[int]]]:
    """
    Align student utterances to key utterances by transcript similarity
    (Needleman-Wunsch style dynamic programming, gaps allowed on both sides).

    Returns a list of (student_index, key_index) pairs in order; None on one
    side means that utterance has no counterpart (student skipped a key
    utterance, or coded an extra one). This replaces the old
    "student_utts[i] vs key_utts[i]" pairing, where a single skipped
    utterance early in the session desynchronized every comparison after it
    and produced a cascade of false MISSING/EXTRA/wrong-modifier flags.
    """
    n, m = len(student_utts), len(key_utts)
    GAP = -0.05  # small penalty so the DP prefers matching over gapping

    # score[i][j] = best total for first i student utts vs first j key utts
    score = [[0.0] * (m + 1) for _ in range(n + 1)]
    move = [[None] * (m + 1) for _ in range(n + 1)]  # 'd'iag, 'u'p (skip student), 'l'eft (skip key)
    for i in range(1, n + 1):
        score[i][0] = score[i - 1][0] + GAP
        move[i][0] = "u"
    for j in range(1, m + 1):
        score[0][j] = score[0][j - 1] + GAP
        move[0][j] = "l"

    sims = [[_text_similarity(
                student_utts[i].utterance_text, key_utts[j].utterance_text,
                a_partial=student_utts[i].text_is_partial,
                b_partial=key_utts[j].text_is_partial)
             for j in range(m)] for i in range(n)]

    for i in range(1, n + 1):
        for j in range(1, m + 1):
            sim = sims[i - 1][j - 1]
            diag = score[i - 1][j - 1] + (sim if sim >= min_similarity else GAP * 2)
            up = score[i - 1][j] + GAP
            left = score[i][j - 1] + GAP
            best = max(diag, up, left)
            score[i][j] = best
            move[i][j] = "d" if best == diag else ("u" if best == up else "l")

    pairs: List[Tuple[Optional[int], Optional[int]]] = []
    i, j = n, m
    while i > 0 or j > 0:
        mv = move[i][j]
        if mv == "d":
            si, kj = i - 1, j - 1
            if sims[si][kj] >= min_similarity:
                pairs.append((si, kj))
            else:
                # diagonal was taken as a double-gap: report both as unmatched
                pairs.append((si, None))
                pairs.append((None, kj))
            i, j = i - 1, j - 1
        elif mv == "u":
            pairs.append((i - 1, None))
            i -= 1
        else:
            pairs.append((None, j - 1))
            j -= 1
    pairs.reverse()
    return pairs


def check_transcript(student_utt: Utterance, key_utt: Utterance, grammar: Dict[str, Any]) -> List[Issue]:
    """Flag a transcript that differs from the key's.

    The codes can all be correct while the transcript itself is wrong (e.g.
    a student transcribed "-S DOG BARN" where the key has "-S BARN DOG" -
    same symbols, wrong order). Reviewers catch this by eye, so the checker
    should too. Comparison ignores case, spacing and bracket/paren notation,
    since relevance-marking is a separate judgment call and is not graded
    here. This is a note-level (orange) issue: it does not affect the 1/0
    Scores sheet, which only covers the 11 coding categories.
    """
    orange = grammar["colors"]["usv_missing_note"]
    s_text, k_text = student_utt.utterance_text, key_utt.utterance_text
    if not s_text or not k_text:
        return []
    if _match_text(s_text) == _match_text(k_text):
        return []

    ci_row = None
    for row in student_utt.rows:
        if behavior_in(row.behavior, grammar["utterance_detection"]["intent_behaviors"]):
            ci_row = row
            break
    if ci_row is None:
        return []

    return [Issue(
        kind="transcript_mismatch",
        color=orange,
        message=render_feedback(grammar, None, "transcript_mismatch",
                                utterance=s_text, key=k_text),
        row_index=ci_row.original_index,
        utterance_id=student_utt.uid,
        expected=k_text,
    )]


def compare_files_with_alignment(
    student_df: pd.DataFrame, key_df: pd.DataFrame, grammar: Dict[str, Any]
) -> Tuple[List[Issue], List[Utterance], List[Utterance], List[Tuple[Optional[int], Optional[int]]]]:
    student_prepared = prepare_dataframe(student_df, grammar)
    key_prepared = prepare_dataframe(key_df, grammar)

    student_utts = segment_utterances(dataframe_to_events(student_prepared, grammar), grammar)
    key_utts = segment_utterances(dataframe_to_events(key_prepared, grammar), grammar)

    pairs = align_utterances(student_utts, key_utts)

    issues: List[Issue] = []
    red = grammar["colors"]["error"]
    purple = grammar["colors"]["boundary_unclear"]

    last_matched_student_row = 0
    for s_idx, k_idx in pairs:
        if s_idx is not None and k_idx is not None:
            issues.extend(compare_utterances(student_utts[s_idx], key_utts[k_idx], grammar))
            issues.extend(check_transcript(student_utts[s_idx], key_utts[k_idx], grammar))
            last_matched_student_row = student_utts[s_idx].last_row_index
        elif k_idx is not None:
            key_utt = key_utts[k_idx]
            issues.append(Issue(
                kind="missing_utterance",
                color=red,
                message=render_feedback(grammar, None, "missing_utterance",
                                        key=key_utt.utterance_text or "no transcript"),
                expected=key_utt.utterance_text,
                insert_after_row_index=last_matched_student_row,
            ))
        else:
            student_utt = student_utts[s_idx]
            issues.append(Issue(
                kind="extra_utterance",
                color=purple,
                message=render_feedback(grammar, None, "extra_utterance",
                                        utterance=student_utt.utterance_text),
                row_index=student_utt.anchor_row_index,
                utterance_id=student_utt.uid,
            ))
            last_matched_student_row = student_utt.last_row_index

    return issues, student_utts, key_utts, pairs


def compare_files(student_df: pd.DataFrame, key_df: pd.DataFrame, grammar: Dict[str, Any]) -> Tuple[List[Issue], List[Utterance], List[Utterance]]:
    issues, student_utts, key_utts, _ = compare_files_with_alignment(student_df, key_df, grammar)
    return issues, student_utts, key_utts


def extract_first_utterances(df: pd.DataFrame, grammar: Dict[str, Any], n: int = 5) -> List[str]:
    prepared = prepare_dataframe(df, grammar)
    utts = segment_utterances(dataframe_to_events(prepared, grammar), grammar)
    texts = []
    for utt in utts:
        text = utt.utterance_text
        if text and not is_ignored_comment(text, grammar):
            texts.append(text)
        if len(texts) >= n:
            break
    return texts


def create_reference_passport(reference_path: str | Path, grammar: Dict[str, Any], display_name: Optional[str] = None, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    df = read_excel_first_sheet_or_named(reference_path, sheet_name=sheet_name)
    first_n = grammar.get("output", {}).get("first_utterances_in_passport", 5)
    first_utts = extract_first_utterances(df, grammar, n=first_n)
    prepared = prepare_dataframe(df, grammar)
    utts = segment_utterances(dataframe_to_events(prepared, grammar), grammar)
    path = Path(reference_path)
    return {
        "id": path.stem,
        "display_name": display_name or path.stem,
        "file_name": path.name,
        "sheet_name": sheet_name,
        "utterance_count": len(utts),
        "first_utterances": first_utts,
    }


def add_reference_to_library(uploaded_path: str | Path, library_dir: str | Path, grammar: Dict[str, Any], display_name: Optional[str] = None, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    library = Path(library_dir)
    library.mkdir(parents=True, exist_ok=True)
    src = Path(uploaded_path)
    dst = library / src.name
    if src.resolve() != dst.resolve():
        shutil.copy2(src, dst)
    passport = create_reference_passport(dst, grammar, display_name=display_name, sheet_name=sheet_name)
    passport_path = dst.with_suffix(".json")
    passport_path.write_text(json.dumps(passport, ensure_ascii=False, indent=2), encoding="utf-8")
    return passport


def index_reference_folder(library_dir: str | Path, grammar: Dict[str, Any]) -> List[str]:
    """Create passports for any .xlsx in the library folder that lacks one.

    The app reads the library from the .json passports, not from the .xlsx
    files themselves. Previously a passport was only written when a file was
    uploaded through the Reference Library page, so key files copied straight
    into reference_keys/ were invisible to the app and had to be re-uploaded.
    This scans the folder and indexes anything new, so dropping files into
    reference_keys/ is enough (uploading through the UI still works too).

    Returns the display names of newly indexed files.
    """
    library = Path(library_dir)
    if not library.exists():
        return []
    newly_indexed: List[str] = []
    for xlsx in sorted(library.glob("*.xlsx")):
        if xlsx.name.startswith("~$"):  # Excel lock file
            continue
        passport_path = xlsx.with_suffix(".json")
        if passport_path.exists():
            continue
        try:
            passport = create_reference_passport(xlsx, grammar)
            passport_path.write_text(json.dumps(passport, ensure_ascii=False, indent=2), encoding="utf-8")
            newly_indexed.append(passport.get("display_name", xlsx.stem))
        except Exception:
            # A malformed/unrelated xlsx in the folder shouldn't break startup.
            continue
    return newly_indexed


def load_reference_passports(library_dir: str | Path) -> List[Dict[str, Any]]:
    library = Path(library_dir)
    if not library.exists():
        return []
    passports = []
    for p in sorted(library.glob("*.json")):
        try:
            passports.append(json.loads(p.read_text(encoding="utf-8")))
        except Exception:
            continue
    return passports


def score_passport_match(student_first: List[str], passport: Dict[str, Any]) -> int:
    ref_first = passport.get("first_utterances", [])
    score = 0
    for a, b in zip(student_first, ref_first):
        if norm(a).lower() == norm(b).lower():
            score += 2
        elif norm(a).lower() in norm(b).lower() or norm(b).lower() in norm(a).lower():
            score += 1
    return score


def suggest_references(student_df: pd.DataFrame, grammar: Dict[str, Any], library_dir: str | Path, top_k: int = 3) -> Tuple[List[str], List[Dict[str, Any]]]:
    first_n = grammar.get("output", {}).get("first_utterances_in_passport", 5)
    student_first = extract_first_utterances(student_df, grammar, n=first_n)
    passports = load_reference_passports(library_dir)
    ranked = sorted(passports, key=lambda p: score_passport_match(student_first, p), reverse=True)
    return student_first, ranked[:top_k]


# Report column label for each grammar sequence group, in report order.
# Matches the lab's manual score-sheet format (see Study_1_Video1_Scores_*):
# one row per KEY utterance, one column per coding category, 1 = student
# matches the key for that category, 0 = any discrepancy (wrong behavior,
# wrong/missing/extra modifiers, or missing code row). An utterance the
# student did not code at all scores 0 in every category.
SCORES_COLUMN_ORDER: List[Tuple[str, str]] = [
    ("Listing", "Listing"),
    ("Communicative Intent", "Communicative intent"),
    ("Imitative", "Imitativeness"),
    ("Independent Aided Utterances", "Independence"),
    ("Number of Symbols", "# of symbols"),
    ("Number of Relevant Symbols", "# of relevant symbols"),
    ("Word Order", "Word order"),
    ("SV", "SV"),
    ("Parts of Speech", "Parts of speech"),
    ("Inflectional Morphemes", "Inflectional morph"),
    ("Grammatical Intent", "Grammatical Intent"),
]


def _issue_group(issue: Issue, student_utt: Utterance, key_utt: Utterance, grammar: Dict[str, Any]) -> Optional[str]:
    """Determine which sequence group an issue belongs to."""
    if issue.kind == "missing" and issue.missing_behavior:
        entry = get_sequence_entry(grammar, issue.missing_behavior)
        return entry["group"] if entry else None
    if issue.row_index is not None:
        for row in student_utt.rows:
            if row.original_index == issue.row_index:
                entry = get_sequence_entry(grammar, row.behavior)
                return entry["group"] if entry else None
    return None


def build_scores_rows(
    student_utts: List[Utterance],
    key_utts: List[Utterance],
    pairs: List[Tuple[Optional[int], Optional[int]]],
    issues: List[Issue],
    grammar: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """One dict per KEY utterance with 1/0 per category, in key order.

    Rows carry a private "_missing" flag (True when the student did not code
    that key utterance at all). Only those rows get highlighted in the Scores
    sheet; individual 0s are left unshaded since the numbers speak for
    themselves and shading every 0 made the sheet hard to read.
    """
    # Issues per student utterance uid, bucketed by group.
    failed_groups_by_uid: Dict[int, set] = {}
    utt_by_uid = {u.uid: u for u in student_utts}
    key_by_student_uid: Dict[int, Utterance] = {}
    for s_idx, k_idx in pairs:
        if s_idx is not None and k_idx is not None:
            key_by_student_uid[student_utts[s_idx].uid] = key_utts[k_idx]

    for issue in issues:
        if issue.utterance_id is None or issue.kind in (
            "boundary_unclear", "usv_missing_note", "extra_utterance",
            "transcript_mismatch", "extra",
        ):
            # Note-level flags: reminders and manual-check prompts, not key
            # mismatches. "extra" means the student coded a category the key
            # does not code for this utterance (e.g. coding beyond the three
            # categories required when Listing is present). It is still shown
            # in the annotated sheet, but it costs no marks: the score sheet
            # reflects only the utterances in the key and the way the key
            # codes them. No penalty marks.
            continue
        s_utt = utt_by_uid.get(issue.utterance_id)
        k_utt = key_by_student_uid.get(issue.utterance_id)
        if s_utt is None or k_utt is None:
            continue
        group = _issue_group(issue, s_utt, k_utt, grammar)
        if group:
            failed_groups_by_uid.setdefault(issue.utterance_id, set()).add(group)

    student_by_key_uid: Dict[int, Utterance] = {}
    for s_idx, k_idx in pairs:
        if s_idx is not None and k_idx is not None:
            student_by_key_uid[key_utts[k_idx].uid] = student_utts[s_idx]

    rows: List[Dict[str, Any]] = []
    for number, key_utt in enumerate(key_utts, start=1):
        time_val = ""
        for r in key_utt.rows:
            if behavior_in(r.behavior, grammar["utterance_detection"]["intent_behaviors"]):
                time_val = clean_text(r.data.get("Time_Relative_hms"))
                break
        row: Dict[str, Any] = {
            "Number": number,
            "Time_Relative_hms": time_val,
            "Utterance": key_utt.utterance_text,
        }
        student_utt = student_by_key_uid.get(key_utt.uid)
        failed = failed_groups_by_uid.get(student_utt.uid, set()) if student_utt else None
        row["_missing"] = student_utt is None
        for group, label in SCORES_COLUMN_ORDER:
            if student_utt is None:
                row[label] = 0  # whole utterance missing from student coding
            else:
                row[label] = 0 if group in failed else 1
        rows.append(row)
    return rows


def write_scores_sheet(wb, scores_rows: List[Dict[str, Any]]) -> None:
    """Append a 'Scores' sheet in the lab's manual score-sheet layout."""
    if "Scores" in wb.sheetnames:
        del wb["Scores"]
    ws = wb.create_sheet("Scores")

    headers = ["Number", "Time_Relative_hms", "Utterance"] + [label for _, label in SCORES_COLUMN_ORDER]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    missing_fill = PatternFill("solid", fgColor="FF9999")
    for r, row in enumerate(scores_rows, start=2):
        is_missing = row.get("_missing", False)
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=col, value=row.get(h))
            if is_missing:
                # Student never coded this key utterance: shade the whole row.
                cell.fill = missing_fill

    n = len(scores_rows)
    if n:
        total_row = n + 3
        pct_row = n + 4
        ws.cell(row=total_row, column=3, value="Total").font = Font(bold=True)
        ws.cell(row=pct_row, column=3, value="Percentage").font = Font(bold=True)
        for col_offset, (_, label) in enumerate(SCORES_COLUMN_ORDER):
            col = 4 + col_offset
            total = sum(r[label] for r in scores_rows)
            ws.cell(row=total_row, column=col, value=total).font = Font(bold=True)
            pct_cell = ws.cell(row=pct_row, column=col, value=total / n)
            pct_cell.number_format = "0.0%"
            pct_cell.font = Font(bold=True)

    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 40
    for col in range(4, 4 + len(SCORES_COLUMN_ORDER)):
        ws.column_dimensions[get_column_letter(col)].width = 14


def write_annotated_excel(
    student_input_path: str | Path,
    output_path: str | Path,
    issues: List[Issue],
    grammar: Dict[str, Any],
    student_utts: Optional[List[Utterance]] = None,
    key_utts: Optional[List[Utterance]] = None,
    alignment: Optional[List[Tuple[Optional[int], Optional[int]]]] = None,
    reviewer_notes: Optional[Dict[int, str]] = None,
) -> None:
    """Create a copy of the student's workbook with highlights and inserted MISSING rows.

    Only the required columns (Time_Relative_hms, Observation, Behavior, Comment)
    plus any Modifier* columns are kept in the output - Observer XT exports
    include many extra columns (Date_Time_Absolute..., Event_Log, Event_Type,
    etc.) that aren't part of the coding scheme and were previously left in
    the annotated file untouched, even though prepare_dataframe() already
    ignores them for comparison purposes.

    If student_utts is provided (compare_files() returns it), every row gets
    an "Utterance #" value and every Review_Notes entry is prefixed with
    "[Utt N: "transcript"]" so a reviewer can tell at a glance which
    utterance an issue belongs to instead of having to reconstruct that by
    hand from row position.
    """
    wb = load_workbook(student_input_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    header_map = {str(v).strip(): i + 1 for i, v in enumerate(header) if v is not None}

    required = grammar["columns"]["required"]
    modifier_prefix = grammar["columns"]["modifier_columns"]["detect_by_prefix"]
    keep_names = set(required) | {
        name for name in header_map if str(name).startswith(modifier_prefix)
    }

    cols_to_delete = sorted(
        (idx for name, idx in header_map.items() if name not in keep_names),
        reverse=True,
    )
    for idx in cols_to_delete:
        ws.delete_cols(idx)

    # Recompute column positions now that unwanted columns are gone.
    header = [cell.value for cell in ws[1]]
    header_map = {str(v).strip(): i + 1 for i, v in enumerate(header) if v is not None}

    behavior_col = header_map.get("Behavior")
    comment_col = header_map.get("Comment")
    if behavior_col is None or comment_col is None:
        raise ValueError("Could not find Behavior/Comment columns in the student file.")

    # Map each original row index to its utterance, used to prefix Review_Notes
    # with "[Utt #N: transcript]" so a reviewer can tell which utterance an
    # issue belongs to without adding a separate column.
    row_to_utt: Dict[int, Utterance] = {}
    utt_by_uid: Dict[int, Utterance] = {}
    if student_utts:
        for utt in student_utts:
            utt_by_uid[utt.uid] = utt
            for r in utt.rows:
                row_to_utt[r.original_index] = utt

    # Add Review_Notes column if absent.
    review_col = header_map.get("Review_Notes")
    if review_col is None:
        review_col = ws.max_column + 1
        ws.cell(row=1, column=review_col).value = "Review_Notes"
        ws.cell(row=1, column=review_col).font = Font(bold=True)

    fills = {
        "red": PatternFill("solid", fgColor=grammar["colors"]["error"]),
        "orange": PatternFill("solid", fgColor=grammar["colors"]["usv_missing_note"]),
        "purple": PatternFill("solid", fgColor=grammar["colors"]["boundary_unclear"]),
    }

    # Map issue color hex back to fill by value.
    def fill_for(issue: Issue) -> PatternFill:
        if issue.color == grammar["colors"]["usv_missing_note"]:
            return fills["orange"]
        if issue.color == grammar["colors"]["boundary_unclear"]:
            return fills["purple"]
        return fills["red"]

    def context_prefix(issue: Issue) -> str:
        if issue.utterance_id is None:
            return ""
        utt = utt_by_uid.get(issue.utterance_id)
        text = utt.utterance_text if utt else ""
        label = f"[Utt #{issue.utterance_id}"
        if text:
            label += f': "{text}"'
        return label + "]"

    def annotate(excel_row: int, issue: Issue, label: str, with_note: bool = True) -> None:
        """Short label in the cell, full feedback in a hover comment.

        The feedback texts are long enough that putting them in the cell made
        the sheet unreadable, so the cell shows only which utterance the issue
        belongs to (or a one-word marker) and the explanation lives in an Excel
        cell comment - the same way a human reviewer leaves notes.
        Several issues on one row are merged into a single comment.
        """
        cell = ws.cell(row=excel_row, column=review_col)
        existing_label = clean_text(cell.value)
        cell.value = existing_label if existing_label else label
        if not with_note:
            # The label itself is the whole message (missing utterances), so a
            # hover note would just repeat it.
            return

        previous = cell.comment.text if cell.comment else ""
        body = (previous + "\n\n" if previous else "") + issue.message
        c = CellComment(body, "Coding Checker")
        c.width = 420
        c.height = max(90, 26 * (body.count("\n") + body.count(". ") + 2))
        cell.comment = c

    # Existing row highlights. DataFrame index 0 => Excel row 2.
    for issue in issues:
        if issue.row_index is None:
            continue
        excel_row = issue.row_index + 2
        f = fill_for(issue)
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col).fill = f
        annotate(excel_row, issue, context_prefix(issue))

    # Insert missing rows bottom-up so row indices remain stable.
    insert_issues = [i for i in issues if i.insert_after_row_index is not None]
    insert_issues.sort(key=lambda x: x.insert_after_row_index or 0, reverse=True)
    for issue in insert_issues:
        insert_after_excel_row = (issue.insert_after_row_index or 0) + 2
        insert_at = insert_after_excel_row + 1
        ws.insert_rows(insert_at)
        for col in range(1, ws.max_column + 1):
            ws.cell(row=insert_at, column=col).fill = fill_for(issue)
        ws.cell(row=insert_at, column=behavior_col).value = issue.missing_behavior or "MISSING"
        ws.cell(row=insert_at, column=behavior_col).font = Font(bold=True)
        # The Comment column is left untouched: the explanation belongs in
        # Review_Notes only, so a missing item is reported once, not twice.
        label = context_prefix(issue)
        if not label:
            missing_text = clean_text(issue.expected)
            label = f'[Missing utterance: "{missing_text}"]' if missing_text else "[Missing utterance]"
        annotate(insert_at, issue, label, with_note=(issue.kind != "missing_utterance"))

    # Basic readability.
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = min(max(ws.column_dimensions[letter].width or 12, 14), 35)
    ws.column_dimensions[get_column_letter(comment_col)].width = 35
    ws.column_dimensions[get_column_letter(review_col)].width = 30

    # Reviewer notes typed in the app. They are appended to the same Excel note
    # as the program's own feedback, on the utterance's anchor row, so the
    # student sees one note per row rather than two parallel sets of comments.
    if reviewer_notes and student_utts:
        for utt in student_utts:
            text = clean_text(reviewer_notes.get(utt.uid))
            if not text:
                continue
            excel_row = utt.anchor_row_index + 2
            cell = ws.cell(row=excel_row, column=review_col)
            if not clean_text(cell.value):
                label = f'[Utt #{utt.uid}'
                if utt.utterance_text:
                    label += f': "{utt.utterance_text}"'
                cell.value = label + "]"
                for col in range(1, ws.max_column + 1):
                    if not ws.cell(row=excel_row, column=col).fill.fill_type:
                        ws.cell(row=excel_row, column=col).fill = fills["orange"]
            previous = cell.comment.text if cell.comment else ""
            body = (previous + "\n\n" if previous else "") + "Reviewer: " + text
            c = CellComment(body, "Coding Checker")
            c.width = 420
            c.height = max(90, 26 * (body.count("\n") + 3))
            cell.comment = c

    # Scores sheet (second tab), mirroring the lab's manual score sheet.
    if student_utts is not None and key_utts is not None and alignment is not None:
        scores_rows = build_scores_rows(student_utts, key_utts, alignment, issues, grammar)
        write_scores_sheet(wb, scores_rows)

    wb.save(output_path)
