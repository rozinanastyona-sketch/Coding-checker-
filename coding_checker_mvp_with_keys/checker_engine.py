from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
import json
import re
import shutil

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


@dataclass
class EventRow:
    original_index: int
    data: Dict[str, Any]
    behavior: str
    modifiers: List[str]
    comment: str


@dataclass
class Utterance:
    uid: int
    rows: List[EventRow] = field(default_factory=list)
    boundary_unclear: bool = False

    @property
    def first_row_index(self) -> int:
        return self.rows[0].original_index if self.rows else -1

    @property
    def last_row_index(self) -> int:
        return self.rows[-1].original_index if self.rows else -1

    @property
    def utterance_text(self) -> str:
        for row in self.rows:
            if row.behavior == "Communicative Intent Present":
                c = clean_text(row.comment)
                if c and not c.upper().startswith("USV:"):
                    return c
        return ""


@dataclass
class Issue:
    kind: str
    color: str
    message: str
    row_index: Optional[int] = None
    insert_after_row_index: Optional[int] = None
    missing_behavior: Optional[str] = None
    expected: Optional[str] = None
    utterance_id: Optional[int] = None


def load_grammar(path: str | Path) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    return str(value).strip()


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def build_alias_map(grammar: Dict[str, Any]) -> Dict[str, str]:
    alias_to_canonical = {}
    for canonical, aliases in grammar.get("aliases", {}).items():
        alias_to_canonical[norm(canonical).lower()] = canonical
        for alias in aliases or []:
            alias_to_canonical[norm(alias).lower()] = canonical
    return alias_to_canonical


def canonicalize(value: Any, alias_map: Dict[str, str]) -> str:
    v = norm(value)
    return alias_map.get(v.lower(), v)


def read_excel_first_sheet_or_named(path_or_file: Any, sheet_name: Optional[str] = None) -> pd.DataFrame:
    if sheet_name is None:
        return pd.read_excel(path_or_file, sheet_name=0)
    return pd.read_excel(path_or_file, sheet_name=sheet_name)


def prepare_dataframe(df: pd.DataFrame, grammar: Dict[str, Any]) -> pd.DataFrame:
    """Keep required columns + all Modifier* columns. Do not assume how many modifiers exist."""
    df = df.copy()
    required = grammar["columns"]["required"]
    modifier_prefix = grammar["columns"]["modifier_columns"]["detect_by_prefix"]

    # Observer files may contain whitespace in column names.
    df.columns = [str(c).strip() for c in df.columns]

    missing_required = [c for c in required if c not in df.columns]
    if missing_required:
        raise ValueError(f"Missing required columns: {missing_required}")

    modifier_cols = [c for c in df.columns if str(c).strip().startswith(modifier_prefix)]
    keep_cols = required + modifier_cols
    return df[keep_cols].copy()


def row_to_event(row: pd.Series, idx: int, grammar: Dict[str, Any], alias_map: Dict[str, str]) -> EventRow:
    modifier_prefix = grammar["columns"]["modifier_columns"]["detect_by_prefix"]
    modifier_cols = [c for c in row.index if str(c).startswith(modifier_prefix)]
    modifiers = []
    for col in modifier_cols:
        val = clean_text(row.get(col))
        if val:
            modifiers.append(canonicalize(val, alias_map))

    behavior = canonicalize(row.get("Behavior"), alias_map)
    comment = clean_text(row.get("Comment"))
    data = {str(k): row.get(k) for k in row.index}
    data["Behavior"] = behavior
    data["Comment"] = comment
    return EventRow(original_index=idx, data=data, behavior=behavior, modifiers=modifiers, comment=comment)


def dataframe_to_events(df: pd.DataFrame, grammar: Dict[str, Any]) -> List[EventRow]:
    alias_map = build_alias_map(grammar)
    return [row_to_event(row, idx, grammar, alias_map) for idx, row in df.iterrows()]


def is_ignored_comment(comment: str, grammar: Dict[str, Any]) -> bool:
    prefixes = grammar["utterance_detection"].get("ignore_comment_prefixes", [])
    c = clean_text(comment)
    return any(c.upper().startswith(p.upper()) for p in prefixes)


def segment_utterances(events: List[EventRow], grammar: Dict[str, Any]) -> List[Utterance]:
    """
    CI-anchored segmentation.

    Important design decision:
    - Utterances are NOT primarily built from Utterance start/stop.
    - Each Communicative Intent Present/Absent row is the semantic anchor of an utterance.
    - Rows from one CI anchor up to the row before the next CI anchor belong to the same utterance.
    - Utterance start/stop rows are kept inside the utterance as context, but are not allowed to break segmentation.
    - Duplicate/orphan start/stop rows are treated as Observer noise, not student errors.

    Why this fixes the "one utterance only / purple first row" problem:
    If Observer exports noisy sequences like start-stop-stop or start-start-stop,
    the old state-machine segmentation can get stuck. CI-based segmentation still finds
    all actual utterances because every utterance should have Communicative Intent.
    """
    intents = set(grammar["utterance_detection"]["intent_behaviors"])
    start_b = grammar["utterance_detection"]["start_behavior"]
    stop_b = grammar["utterance_detection"]["stop_behavior"]

    # Find all Communicative Intent rows. These are the utterance anchors.
    ci_positions = [i for i, ev in enumerate(events) if ev.behavior in intents]

    # If there are no CI rows, we cannot reliably segment. Return one purple review chunk.
    if not ci_positions:
        return [Utterance(uid=1, rows=events, boundary_unclear=True)] if events else []

    utterances: List[Utterance] = []

    for uid, ci_pos in enumerate(ci_positions, start=1):
        next_ci_pos = ci_positions[uid] if uid < len(ci_positions) else len(events)

        # Include nearby boundary rows immediately before CI, especially Utterance start.
        # We walk backwards over boundary-only rows so the start row stays with the utterance.
        start_pos = ci_pos
        j = ci_pos - 1
        while j >= 0 and events[j].behavior in {start_b, stop_b}:
            start_pos = j
            j -= 1

        # Do not let this utterance steal rows that belong to the previous CI interval.
        if utterances:
            prev_last = utterances[-1].last_row_index
            # original_index follows DataFrame row order; convert by positions instead below.
            prev_end_pos = max(k for k, ev in enumerate(events) if ev.original_index == prev_last)
            start_pos = max(start_pos, prev_end_pos + 1)

        rows = events[start_pos:next_ci_pos]

        boundary_unclear = False
        # If a CI Present row has no actual utterance text, mark for review.
        ci_event = events[ci_pos]
        if ci_event.behavior == "Communicative Intent Present":
            c = clean_text(ci_event.comment)
            if not c or is_ignored_comment(c, grammar):
                boundary_unclear = True

        utterances.append(Utterance(uid=uid, rows=rows, boundary_unclear=boundary_unclear))

    return utterances

def get_sequence_entry(grammar: Dict[str, Any], behavior: str) -> Optional[Dict[str, Any]]:
    for entry in grammar["sequence"]:
        if behavior in entry.get("behaviors", []):
            return entry
    return None


def comparable_rows(utt: Utterance, grammar: Dict[str, Any]) -> List[EventRow]:
    start_b = grammar["utterance_detection"]["start_behavior"]
    stop_b = grammar["utterance_detection"]["stop_behavior"]
    return [r for r in utt.rows if r.behavior not in {start_b, stop_b}]


def find_row_by_group(utt: Utterance, entry: Dict[str, Any], grammar: Dict[str, Any]) -> Optional[EventRow]:
    behaviors = set(entry.get("behaviors", []))
    for row in comparable_rows(utt, grammar):
        if row.behavior in behaviors:
            return row
    return None


def modifier_set(row: Optional[EventRow]) -> set[str]:
    if row is None:
        return set()
    return {norm(m) for m in row.modifiers if norm(m)}


def validate_row_against_key(student_row: EventRow, key_row: EventRow, entry: Dict[str, Any]) -> Optional[str]:
    if student_row.behavior != key_row.behavior:
        return f"Expected Behavior: {key_row.behavior}; found: {student_row.behavior}"

    expected_mods = modifier_set(key_row)
    student_mods = modifier_set(student_row)
    if expected_mods != student_mods:
        missing = sorted(expected_mods - student_mods)
        extra = sorted(student_mods - expected_mods)
        parts = []
        if missing:
            parts.append("Missing modifiers: " + ", ".join(missing))
        if extra:
            parts.append("Extra modifiers: " + ", ".join(extra))
        return "; ".join(parts)
    return None


def check_usv_rule(student_utt: Utterance, grammar: Dict[str, Any]) -> Optional[Issue]:
    for rule in grammar.get("special_rules", []):
        if rule.get("name") != "USV comment required":
            continue
        for row in comparable_rows(student_utt, grammar):
            if row.behavior != rule.get("if_behavior"):
                continue
            mods = modifier_set(row)
            required = set(rule.get("required_modifiers", []))
            if required.issubset(mods):
                prefix = rule.get("then_comment_must_start_with", "USV:")
                # USV note may be in the same row or in any row inside the same utterance.
                has_usv = any(clean_text(r.comment).upper().startswith(prefix.upper()) for r in student_utt.rows)
                if not has_usv:
                    return Issue(
                        kind="usv_missing_note",
                        color=grammar["colors"]["usv_missing_note"],
                        message=f"SV is Lexical + Unique, but Comment does not contain '{prefix}'",
                        row_index=row.original_index,
                        utterance_id=student_utt.uid,
                    )
    return None


def compare_utterances(student_utt: Utterance, key_utt: Utterance, grammar: Dict[str, Any]) -> List[Issue]:
    issues: List[Issue] = []
    red = grammar["colors"]["error"]
    purple = grammar["colors"]["boundary_unclear"]

    if student_utt.boundary_unclear:
        issues.append(Issue(
            kind="boundary_unclear",
            color=purple,
            message="Utterance boundary could not be confidently detected; check manually.",
            row_index=student_utt.first_row_index,
            utterance_id=student_utt.uid,
        ))

    previous_student_row: Optional[EventRow] = None
    for entry in grammar["sequence"]:
        key_row = find_row_by_group(key_utt, entry, grammar)
        student_row = find_row_by_group(student_utt, entry, grammar)

        if key_row is None:
            # Reference does not require this group for this utterance.
            continue

        if student_row is None:
            insert_after = previous_student_row.original_index if previous_student_row else student_utt.first_row_index
            expected = key_row.behavior
            if key_row.modifiers:
                expected += " | " + " | ".join(key_row.modifiers)
            issues.append(Issue(
                kind="missing",
                color=red,
                message=f"MISSING: {expected}",
                insert_after_row_index=insert_after,
                missing_behavior=key_row.behavior,
                expected=expected,
                utterance_id=student_utt.uid,
            ))
            continue

        msg = validate_row_against_key(student_row, key_row, entry)
        if msg:
            issues.append(Issue(
                kind="wrong",
                color=red,
                message=msg,
                row_index=student_row.original_index,
                utterance_id=student_utt.uid,
            ))

        previous_student_row = student_row

    # Extra rows: student has known code groups that key does not have in that utterance.
    for srow in comparable_rows(student_utt, grammar):
        entry = get_sequence_entry(grammar, srow.behavior)
        if entry is None:
            continue
        krow = find_row_by_group(key_utt, entry, grammar)
        if krow is None:
            issues.append(Issue(
                kind="extra",
                color=red,
                message=f"EXTRA code not expected in key utterance: {srow.behavior}",
                row_index=srow.original_index,
                utterance_id=student_utt.uid,
            ))

    usv_issue = check_usv_rule(student_utt, grammar)
    if usv_issue:
        issues.append(usv_issue)

    return issues


def compare_files(student_df: pd.DataFrame, key_df: pd.DataFrame, grammar: Dict[str, Any]) -> Tuple[List[Issue], List[Utterance], List[Utterance]]:
    student_prepared = prepare_dataframe(student_df, grammar)
    key_prepared = prepare_dataframe(key_df, grammar)

    student_utts = segment_utterances(dataframe_to_events(student_prepared, grammar), grammar)
    key_utts = segment_utterances(dataframe_to_events(key_prepared, grammar), grammar)

    issues: List[Issue] = []
    n = min(len(student_utts), len(key_utts))
    for i in range(n):
        issues.extend(compare_utterances(student_utts[i], key_utts[i], grammar))

    red = grammar["colors"]["error"]
    purple = grammar["colors"]["boundary_unclear"]
    if len(student_utts) < len(key_utts):
        for key_utt in key_utts[len(student_utts):]:
            issues.append(Issue(
                kind="missing_utterance",
                color=red,
                message=f"MISSING whole utterance from key: {key_utt.utterance_text or 'no text'}",
                insert_after_row_index=student_utts[-1].last_row_index if student_utts else 0,
            ))
    elif len(student_utts) > len(key_utts):
        for student_utt in student_utts[len(key_utts):]:
            issues.append(Issue(
                kind="extra_utterance",
                color=purple,
                message="Extra student utterance or mismatch in utterance boundaries; check manually.",
                row_index=student_utt.first_row_index,
                utterance_id=student_utt.uid,
            ))

    return issues, student_utts, key_utts


def extract_first_utterances(df: pd.DataFrame, grammar: Dict[str, Any], n: int = 5) -> List[str]:
    prepared = prepare_dataframe(df, grammar)
    utts = segment_utterances(dataframe_to_events(prepared, grammar), grammar)
    texts = []
    for utt in utts:
        text = utt.utterance_text
        if text and not is_ignored_comment(text, grammar):
            texts.append(text)
        if len(texts) >= n:
            break
    return texts


def create_reference_passport(reference_path: str | Path, grammar: Dict[str, Any], display_name: Optional[str] = None, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    df = read_excel_first_sheet_or_named(reference_path, sheet_name=sheet_name)
    first_n = grammar.get("output", {}).get("first_utterances_in_passport", 5)
    first_utts = extract_first_utterances(df, grammar, n=first_n)
    prepared = prepare_dataframe(df, grammar)
    utts = segment_utterances(dataframe_to_events(prepared, grammar), grammar)
    path = Path(reference_path)
    return {
        "id": path.stem,
        "display_name": display_name or path.stem,
        "file_name": path.name,
        "sheet_name": sheet_name,
        "utterance_count": len(utts),
        "first_utterances": first_utts,
    }


def add_reference_to_library(uploaded_path: str | Path, library_dir: str | Path, grammar: Dict[str, Any], display_name: Optional[str] = None, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    library = Path(library_dir)
    library.mkdir(parents=True, exist_ok=True)
    src = Path(uploaded_path)
    dst = library / src.name
    if src.resolve() != dst.resolve():
        shutil.copy2(src, dst)
    passport = create_reference_passport(dst, grammar, display_name=display_name, sheet_name=sheet_name)
    passport_path = dst.with_suffix(".json")
    passport_path.write_text(json.dumps(passport, ensure_ascii=False, indent=2), encoding="utf-8")
    return passport


def load_reference_passports(library_dir: str | Path) -> List[Dict[str, Any]]:
    library = Path(library_dir)
    if not library.exists():
        return []
    passports = []
    for p in sorted(library.glob("*.json")):
        try:
            passports.append(json.loads(p.read_text(encoding="utf-8")))
        except Exception:
            continue
    return passports


def score_passport_match(student_first: List[str], passport: Dict[str, Any]) -> int:
    ref_first = passport.get("first_utterances", [])
    score = 0
    for a, b in zip(student_first, ref_first):
        if norm(a).lower() == norm(b).lower():
            score += 2
        elif norm(a).lower() in norm(b).lower() or norm(b).lower() in norm(a).lower():
            score += 1
    return score


def suggest_references(student_df: pd.DataFrame, grammar: Dict[str, Any], library_dir: str | Path, top_k: int = 3) -> Tuple[List[str], List[Dict[str, Any]]]:
    first_n = grammar.get("output", {}).get("first_utterances_in_passport", 5)
    student_first = extract_first_utterances(student_df, grammar, n=first_n)
    passports = load_reference_passports(library_dir)
    ranked = sorted(passports, key=lambda p: score_passport_match(student_first, p), reverse=True)
    return student_first, ranked[:top_k]


def write_annotated_excel(student_input_path: str | Path, output_path: str | Path, issues: List[Issue], grammar: Dict[str, Any]) -> None:
    """Create a copy of the student's workbook with highlights and inserted MISSING rows."""
    wb = load_workbook(student_input_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    header_map = {str(v).strip(): i + 1 for i, v in enumerate(header) if v is not None}

    behavior_col = header_map.get("Behavior")
    comment_col = header_map.get("Comment")
    if behavior_col is None or comment_col is None:
        raise ValueError("Could not find Behavior/Comment columns in the student file.")

    # Add Review_Notes column if absent.
    review_col = header_map.get("Review_Notes")
    if review_col is None:
        review_col = ws.max_column + 1
        ws.cell(row=1, column=review_col).value = "Review_Notes"
        ws.cell(row=1, column=review_col).font = Font(bold=True)

    fills = {
        "red": PatternFill("solid", fgColor=grammar["colors"]["error"]),
        "orange": PatternFill("solid", fgColor=grammar["colors"]["usv_missing_note"]),
        "purple": PatternFill("solid", fgColor=grammar["colors"]["boundary_unclear"]),
    }

    # Map issue color hex back to fill by value.
    def fill_for(issue: Issue) -> PatternFill:
        if issue.color == grammar["colors"]["usv_missing_note"]:
            return fills["orange"]
        if issue.color == grammar["colors"]["boundary_unclear"]:
            return fills["purple"]
        return fills["red"]

    # Existing row highlights. DataFrame index 0 => Excel row 2.
    for issue in issues:
        if issue.row_index is None:
            continue
        excel_row = issue.row_index + 2
        f = fill_for(issue)
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col).fill = f
        old = clean_text(ws.cell(row=excel_row, column=review_col).value)
        ws.cell(row=excel_row, column=review_col).value = (old + " | " if old else "") + issue.message

    # Insert missing rows bottom-up so row indices remain stable.
    insert_issues = [i for i in issues if i.insert_after_row_index is not None]
    insert_issues.sort(key=lambda x: x.insert_after_row_index or 0, reverse=True)
    for issue in insert_issues:
        insert_after_excel_row = (issue.insert_after_row_index or 0) + 2
        insert_at = insert_after_excel_row + 1
        ws.insert_rows(insert_at)
        for col in range(1, ws.max_column + 1):
            ws.cell(row=insert_at, column=col).fill = fill_for(issue)
        ws.cell(row=insert_at, column=behavior_col).value = issue.missing_behavior or "MISSING"
        ws.cell(row=insert_at, column=comment_col).value = issue.expected or issue.message
        ws.cell(row=insert_at, column=review_col).value = issue.message
        ws.cell(row=insert_at, column=behavior_col).font = Font(bold=True)

    # Basic readability.
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = min(max(ws.column_dimensions[letter].width or 12, 14), 35)
    ws.column_dimensions[get_column_letter(comment_col)].width = 35
    ws.column_dimensions[get_column_letter(review_col)].width = 45

    wb.save(output_path)
